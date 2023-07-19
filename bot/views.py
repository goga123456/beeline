import datetime
import logging
import time
from datetime import datetime
from threading import Thread

import telebot
from django.http import JsonResponse
from django.views import View
from openpyxl import load_workbook
from telebot import types
from telebot.storage import StateMemoryStorage

from core.settings import BOT_TOKEN, BOT_URL, FILE_EXCEL
import os
import re

logger = telebot.logger
telebot.logger.setLevel(logging.DEBUG)

state_storage = StateMemoryStorage()

bot = telebot.TeleBot(BOT_TOKEN)


class BotAPIView(View):
    def post(self, request):
        json_string = request.body.decode('UTF-8')
        print(123)
        update = telebot.types.Update.de_json(json_string)
        bot.process_new_updates([update])
        return JsonResponse({'code': 200})


user_dict = {}
current_shown_dates = {}
filename = FILE_EXCEL

os.system(f"chmod +x {filename}")

lang_dict = {'wrong_data': {'Русский 🇷🇺': 'Неверные данные', 'Oʻzbek tili 🇺🇿': 'Notoʻgʻri maʻlumotlar'},
             'ask_name': {'Русский 🇷🇺': 'Напиши своё имя', 'Oʻzbek tili 🇺🇿': 'Ismingizni yozing'},
             'ask_surname': {'Русский 🇷🇺': 'Напиши свою фамилию', 'Oʻzbek tili 🇺🇿': 'Familiyangizni yozing'},
             'wrong_name': {'Русский 🇷🇺': 'Данные введены некорректно. Просим указать имя',
                            'Oʻzbek tili 🇺🇿': 'Maʻlumotlar notoʻgʻri kiritilgan. Iltimos, ismni koʻrsating'},
             'wrong_surname': {'Русский 🇷🇺': 'Данные введены некорректно. Просим указать фамилию',
                               'Oʻzbek tili 🇺🇿': 'Maʻlumotlar notoʻgʻri kiritilgan. Iltimos, familiyani koʻrsating'},
             'ask_birthday': {'Русский 🇷🇺': 'Дата твоего рождения',
                              'Oʻzbek tili 🇺🇿': 'Tugʻilgan kun, oy va yilingizni kiriting'},
             'wrong_birthday': {'Русский 🇷🇺': 'Вы ввели неправильную дату!',
                                'Oʻzbek tili 🇺🇿': 'Siz notoʻgʻri sanani kiritdingiz!'},
             'number': {'Русский 🇷🇺': 'Укажи контактный номер, чтобы мы могли связаться с тобой',
                        'Oʻzbek tili 🇺🇿': 'Siz bilan bogʻlanishimiz uchun telefon raqamingizni kiriting'},
             'wrong_number': {'Русский 🇷🇺': 'Неверный формат номера!',
                              'Oʻzbek tili 🇺🇿': 'Notoʻgʻri raqam formati!'},
             'adress': {'Русский 🇷🇺': 'Укажи адрес проживания', 'Oʻzbek tili 🇺🇿': 'Yashash manzilingizni kiriting'},
             'town': {'Русский 🇷🇺': 'Откуда ты?', 'Oʻzbek tili 🇺🇿': 'Qayerdansiz?'},
             'wrong_town_and_district': {
                 'Русский 🇷🇺': 'Название города должно состоять из букв и может быть несколькими словами',
                 'Oʻzbek tili 🇺🇿': 'Shahar nomi harflardan iborat boʻlishi kerak va bir necha so‘z boʻlishi mumkin'},
             'district': {'Русский 🇷🇺': 'Выбери район', 'Oʻzbek tili 🇺🇿': 'Tumanni tanlang'},
             'town_and_districtt': {'Русский 🇷🇺': 'Напиши регион (город) и район',
                                    'Oʻzbek tili 🇺🇿': 'Viloyat (shahar) va tuman nomini yozing'},
             'quarter': {'Русский 🇷🇺': 'Квартал или улица:',
                         'Oʻzbek tili 🇺🇿': 'Kvartal raqami yoki ko‘chaning nomi:'},
             'wrong_quarter': {'Русский 🇷🇺': 'Название квартала или улицы должно состоять из букв или цифр',
                               'Oʻzbek tili 🇺🇿': 'Blok yoki ko‘chaning nomi harflar yoki raqamlardan iborat boʻlishi kerak'},
             'house': {'Русский 🇷🇺': 'Дом:', 'Oʻzbek tili 🇺🇿': 'Uy raqami:'},
             'wrong_house': {'Русский 🇷🇺': 'Название дома должно состоять из цифр или букв',
                             'Oʻzbek tili 🇺🇿': 'Uyning nomi raqamlar yoki harflardan iborat boʻlishi kerak'},
             'education': {'Русский 🇷🇺': 'Укажи уровень образования',
                           'Oʻzbek tili 🇺🇿': 'Taʻlim darajasini ko‘rsating'},
             'uzb_language': {'Русский 🇷🇺': 'Уровень владения Узбекским языком',
                              'Oʻzbek tili 🇺🇿': 'Oʻzbek tilini bilish darajasi'},
             'rus_language': {'Русский 🇷🇺': 'Уровень владения Русским языком',
                              'Oʻzbek tili 🇺🇿': 'Rus tilini bilish darajasi'},
             'eng_language': {'Русский 🇷🇺': 'Уровень владения Английским языком',
                              'Oʻzbek tili 🇺🇿': 'Ingliz tilini bilish darajasi'},
             'higher': {'Русский 🇷🇺': 'Высшее', 'Oʻzbek tili 🇺🇿': 'Oliy'},
             'incomplete_higher': {'Русский 🇷🇺': 'Неполное высшее', 'Oʻzbek tili 🇺🇿': 'Tugallanmagan oliy'},
             'info': {'Русский 🇷🇺': 'На следущие вопросы ответьте выбором одного из вариантов',
                      'Oʻzbek tili 🇺🇿': 'Quyidagi savollarga variantlardan birini tanlash bilan javob bering'},
             'secondary': {'Русский 🇷🇺': 'Среднее', 'Oʻzbek tili 🇺🇿': 'Oʻrta'},
             'incomplete_secondary': {'Русский 🇷🇺': 'Неполное среднее', 'Oʻzbek tili 🇺🇿': 'Tugallanmagan oʻrta'},
             'secondary_special': {'Русский 🇷🇺': 'Среднее специальное', 'Oʻzbek tili 🇺🇿': 'Oʻrta maxsus'},
             'great': {'Русский 🇷🇺': 'Владею как родным языком :)', 'Oʻzbek tili 🇺🇿': 'Men ona tilimda gapiraman :)'},
             'good': {'Русский 🇷🇺': 'Могу вести обычные беседы свободно и уверенно', 'Oʻzbek tili 🇺🇿': 'Oddiy suhbatlarni erkin va ishonchli davom ettirishim mumkin'},
             'satisfactorily': {'Русский 🇷🇺': 'Понимаю язык, но не могу свободно общаться', 'Oʻzbek tili 🇺🇿': 'Men tilni tushunaman, lekin ravon gaplasha olmayman'},
             'ne_vladeyu': {'Русский 🇷🇺': 'Не владею языком', 'Oʻzbek tili 🇺🇿': 'Men ushbu tilda gapirmayman'},
             'work': {'Русский 🇷🇺': 'Есть ли у тебя опыт работы? (неважно официальный или неофициальный)',
                      'Oʻzbek tili 🇺🇿': 'Siz oldin ishlaganmisiz? (rasmiy yoki norasmiy boʻlishidan qatʻiy nazar)'},
             'work_experience': {
                 'Русский 🇷🇺': 'Опиши последний опыт работы коротко\n\n- Кем ты работал? \n- В какой организации? \n- Период работы ',
                 'Oʻzbek tili 🇺🇿': 'Oxirgi ish joyingiz haqida qisqacha maʻlumot bering\n\n- Qaysi lavozimda ishlagansiz? \n- Qaysi tashkilotda? \n- Ishlagan vaqtingiz'},
             'organization': {'Русский 🇷🇺': 'Укажите название организации:',
                              'Oʻzbek tili 🇺🇿': 'Tashkilot nomini kiriting'},
             'wrong_organization': {
                 'Русский 🇷🇺': 'Название организации должно состоять из букв или цифр и может быть несколькими словами',
                 'Oʻzbek tili 🇺🇿': 'Tashkilot nomi harflar yoki raqamlardan iborat boʻlishi kerak va bir nechta soʻzlar boʻlishi mumkin'},
             'job_title': {'Русский 🇷🇺': 'Должность:', 'Oʻzbek tili 🇺🇿': 'Lavozim:'},
             'wrong_job_title': {
                 'Русский 🇷🇺': 'Название специальности должно состоять из букв, также в нём могут быть пробелы и цифры',
                 'Oʻzbek tili 🇺🇿': 'Mutaxassislikning nomi harflardan iborat boʻlishi kerak, unda bo‘shliqlar va raqamlar ham boʻlishi mumkin'},
             'work_start': {'Русский 🇷🇺': 'Укажите год трудоустройства в организацию:',
                            'Oʻzbek tili 🇺🇿': 'Tashkilotga ishga kirgan yilingizni kiriting:'},
             'wrong_work_start': {'Русский 🇷🇺': 'Формат года указан не верно.\nПример: 2020',
                                  'Oʻzbek tili 🇺🇿': 'Yil kiritilgan format noto‘g‘ri.\nMisol: 2020'},
             'work_end': {'Русский 🇷🇺': 'Укажите год увольнения из организации:',
                          'Oʻzbek tili 🇺🇿': 'Siz tashkilotdan boʻshagan yilni koʻrsating:'},
             'wrong_work_end': {'Русский 🇷🇺': 'Формат года указан не верно.\nПример: 2020',
                                'Oʻzbek tili 🇺🇿': 'Yil kiritilgan format noto‘g‘ri.\nMisol: 2020'},
             'wrong_work_datas': {
                 'Русский 🇷🇺': ' Вы не могли уйти с работы раньше чем на неё устроились.Год когда вы устроились на работу?',
                 'Oʻzbek tili 🇺🇿': 'Siz tashkilotdan boʻshagan yilingiz - ishga kirgan yilingizdan oldin boʻlishi mumkin emas. Siz tashkilotga ishga kirgan yilni qaytadan kiriting:'},
             'thank_you': {'Русский 🇷🇺': 'Супер!👍\nСпасибо за прохождение опроса!\n\nТеперь наш рекрутер оценит твой профиль и созвонится с тобой для проведения первого интервью!',
                           'Oʻzbek tili 🇺🇿': 'Zoʻr!!👍\nSoʻrovnomada qatnashganingiz uchun rahmat!\n\nEndi biz Sizning profilingizni baholab birinchi suhbatdan oʻtkazish uchun bogʻlanamiz'},
             'sendmail': {
                 'Русский 🇷🇺': 'Несколько советов  к телефонному интервью:\n\nУбедись, что тебе ничего не будет мешать и ты не будешь отвлекаться на посторонние дела во время телефонного интервью.\n\nЕсли тебе будет неудобно говорить - не стесняйся попросить перенести звонок.\n\nСписок примерных вопросов:\n1. Коротко расскажи о себе\n2. Подробнее расскажи о причинах того, почему ты хочешь работать в Билайн сотрудником Контакт-центра\n3. Что ты ожидаешь получить от работы в Билайн\n\nУдачи :)',
                 'Oʻzbek tili 🇺🇿': 'Telefon orqali suhbatlashish uchun bir nechta maslahatlar:\n\n Hech narsa sizga xalaqit bermasligiga va telefon orqali suhbat davomida sizni begona narsalar chalgʻitmasligiga ishonch hosil qiling. \n\n Agar gaplashayotganda oʻzingizni noqulay his qilsangiz, qoʻngʻiroqni qayta rejalashtirishni soʻrang.\n\nNamuna savollar roʻyxati: \n1. Oʻzingiz haqingizda qisqacha gapirib bering\n2. Nima uchun “Beeline”da Kontakt Markazi xodimi sifatida ishlashni xohlayotganingiz sabablari haqida ko‘proq gapirib bering.\n 3. “Bilayn”da ishlashdan nimani kutmoqdasiz\n\nOmad :)'},
             'again': {'Русский 🇷🇺': 'Если хочешь пройти опрос заново нажми на кнопку: "/start" ',
                       'Oʻzbek tili 🇺🇿': 'Soʻrovnomadan qaytadan oʻtishni istasangiz quyidagi tugmani bosing: "/start"'},
             'checker': {'Русский 🇷🇺': 'Выбери вариант кнопкой',
                         'Oʻzbek tili 🇺🇿': 'Tugmani bosib variantni tanlang'},
             'yes': {'Русский 🇷🇺': 'Ecть', 'Oʻzbek tili 🇺🇿': 'Ha'},
             'no': {'Русский 🇷🇺': 'Нет', 'Oʻzbek tili 🇺🇿': 'Yoʻq'},
             'back': {'Русский 🇷🇺': 'Назад', 'Oʻzbek tili 🇺🇿': 'Ortga'},
             'start': {'Русский 🇷🇺': 'Начать сначала', 'Oʻzbek tili 🇺🇿': 'Boshidan boshlash'},
             'knopka': {'Русский 🇷🇺': 'На следующие вопросы ответь выбором одного из вариантов!',
                        'Oʻzbek tili 🇺🇿': 'Quyidagi savollarga keltirilgan variantlardan birini tanlash orqali javob bering!'},
             'wrong_work_experience': {'Русский 🇷🇺': 'Неверные данные', 'Oʻzbek tili 🇺🇿': 'Notoʻgʻri maʻlumotlar'},
             'tashkent': {'Русский 🇷🇺': 'Ташкент', 'Oʻzbek tili 🇺🇿': 'Toshkent'},
             'drugoi': {'Русский 🇷🇺': 'Другой город или регион', 'Oʻzbek tili 🇺🇿': 'Boshqa viloyat yoki shahar'},
             'Olmazor': {'Русский 🇷🇺': 'Алмазар', 'Oʻzbek tili 🇺🇿': 'Olmazor'},
             'Bektemir': {'Русский 🇷🇺': 'Бектемир', 'Oʻzbek tili 🇺🇿': 'Bektemir'},
             'Mirabad': {'Русский 🇷🇺': 'Мирабад', 'Oʻzbek tili 🇺🇿': 'Mirobod'},
             'Mirzo_Ulugbek': {'Русский 🇷🇺': 'Мирзо-Улугбек', 'Oʻzbek tili 🇺🇿': 'Mirzo Ulugʻbek'},
             'Sergeli': {'Русский 🇷🇺': 'Сергели', 'Oʻzbek tili 🇺🇿': 'Sirgʻali'},
             'Chilonzor': {'Русский 🇷🇺': 'Чиланзар', 'Oʻzbek tili 🇺🇿': 'Chilonzor'},
             'Shayhontohur': {'Русский 🇷🇺': 'Шайхантаур', 'Oʻzbek tili 🇺🇿': 'Shayxontohur'},
             'Yunusobod': {'Русский 🇷🇺': 'Юнусабад', 'Oʻzbek tili 🇺🇿': 'Yunusobod'},
             'Yakkosoroy': {'Русский 🇷🇺': 'Яккасарай', 'Oʻzbek tili 🇺🇿': 'Yakkasoroy'},
             'Yashnobod': {'Русский 🇷🇺': 'Яшнабад', 'Oʻzbek tili 🇺🇿': 'Yashnobod'},
             'Uchtepa': {'Русский 🇷🇺': 'Учтепа', 'Oʻzbek tili 🇺🇿': 'Uchtepa'},
             'Yengihayot': {'Русский 🇷🇺': 'Янгихаёт', 'Oʻzbek tili 🇺🇿': 'Yangihayot'},             
             'prodoljit': {'Русский 🇷🇺': 'Продолжить', 'Oʻzbek tili 🇺🇿': 'Davom etish'},
             'otkazatsya': {'Русский 🇷🇺': 'Отказаться', 'Oʻzbek tili 🇺🇿': 'Rad etish'},
             'want_work_in_bilain': {'Русский 🇷🇺': 'Да, я хочу в Билайн!',
                                     'Oʻzbek tili 🇺🇿': 'Ha, men Beeline da ishlashni xohlayman!'},
             'ne_interesuyet': {'Русский 🇷🇺': 'Не интересует', 'Oʻzbek tili 🇺🇿': 'Qiziqtirmaydi'},
             'resume_text_full': {
                 'Русский 🇷🇺': 'Давай ещё раз уточним, что тебе предстоит пройти:\n\n1. Ты будешь рассматриваться на вакансию оператора контакт-центра Билайн\n\n2. Перед тем как устроиться на работу к нам тебе предстоит пройти 3 этапа отбора: \n- телефонное собеседование\n- тестирование в нашем офисе (там ничего сложного, даже весело 🙂)\n- прикольное собеседование с 2-3 веселыми людьми\n\n3. Если проходишь все этапы - то мы тебя зачисляем на обучение в нашем офисе. \nОбучение длится 15-17 дней. Ты можешь выбрать дневную или вечернюю форму обучения.\nЕсли проходишь обучение и сдаёшь аттестацию - ты принят в штат! 🎉😁\n\nОплата труда, график работы и команда - обо всем этом расскажем и даже покажем 😎\n\nЕсли ты умеешь общаться как минимум на 2 языках (узбекский и русский) и тебе уже 18 лет, жми «Да, я хочу в Билайн!»',
                 'Oʻzbek tili 🇺🇿': 'Keling, tanlov haqida sizga batafsil maʻlumotlar berib oʻtaman:\n\n1. Siz “Beeline” aloqa markazi operatori lavozimiga ko‘rib chiqilasiz\n\n2. Bizga ishga kirishdan oldin siz 3ta bosqichdan o‘tishingiz kerak bo‘ladi: \n- telefon orqali suhbat\n- ofisimizda bir nechta testlar (murakkab narsa yo‘q, aksincha qiziqarli 🙂)\n- 2-3ta xushchaqchaq odamlar bilan ajoyib suhbat\n\n3. Agar siz barcha bosqichlardan muvaffaqiyatli o‘tsangiz, biz sizni o‘quv jarayoniga qabul qilamiz. \nO‘quv jarayoni 15-17 kun davom etadi. Siz kunduzgi yoki kechki guruhni tanlashingiz mumkin.\nO‘quv jarayonini to‘liq o‘qib bo‘lib attestatsiyadan o‘tsangiz - sizni shtatga qabul qilamiz! 🎉😁\n\nIsh haqi, ish jadvali va jamoa - bularning barchasi haqida sizga aytib beramiz va hatto ko‘rsatamiz 😎\n\nAgar siz kamida 2 tilda (o‘zbek va rus) muloqot qila olsangiz va 18 yoshda bo‘lsangiz, unda "Ha, men Beeline da ishlashni xohlayman" tugmasini bosing.'},
             'resume_text': {
                 'Русский 🇷🇺': '1. Ты будешь рассматриваться на вакансию оператора контакт-центра Билайн\n\n2. Перед тем как устроиться на работу к нам тебе предстоит пройти 3 этапа отбора: \n- телефонное собеседование\n- тестирование в нашем офисе (там ничего сложного, даже весело 🙂)\n- прикольное собеседование с 2-3 веселыми людьми\n\n3. Если проходишь все этапы - то мы тебя зачисляем на обучение в нашем офисе. \nОбучение длится 15-17 дней. Ты можешь выбрать дневную или вечернюю форму обучения.\nЕсли проходишь обучение и сдаёшь аттестацию - ты принят в штат! 🎉😁\n\nОплата труда, график работы и команда - обо всем этом расскажем и даже покажем 😎\n\nЕсли ты умеешь общаться как минимум на 2 языках (узбекский и русский) и тебе уже 18 лет, жми «Да, я хочу в Билайн!»',
                 'Oʻzbek tili 🇺🇿': '1. Siz “Beeline” aloqa markazi operatori lavozimiga ko‘rib chiqilasiz\n\n2. Bizga ishga kirishdan oldin siz 3ta bosqichdan o‘tishingiz kerak bo‘ladi: \n- telefon orqali suhbat\n- ofisimizda bir nechta testlar (murakkab narsa yo‘q, aksincha qiziqarli 🙂)\n- 2-3ta xushchaqchaq odamlar bilan ajoyib suhbat\n\n3. Agar siz barcha bosqichlardan muvaffaqiyatli o‘tsangiz, biz sizni o‘quv jarayoniga qabul qilamiz. \nO‘quv jarayoni 15-17 kun davom etadi. Siz kunduzgi yoki kechki guruhni tanlashingiz mumkin.\nO‘quv jarayonini to‘liq o‘qib bo‘lib attestatsiyadan o‘tsangiz - sizni shtatga qabul qilamiz! 🎉😁\n\nIsh haqi, ish jadvali va jamoa - bularning barchasi haqida sizga aytib beramiz va hatto ko‘rsatamiz 😎\n\nAgar siz kamida 2 tilda (o‘zbek va rus) muloqot qila olsangiz va 18 yoshda bo‘lsangiz, unda "Ha, men Beeline da ishlashni xohlayman" tugmasini bosing.'},
             'resume_text_start': {'Русский 🇷🇺': 'Давай ещё раз уточним, что тебе предстоит пройти:',
                                   'Oʻzbek tili 🇺🇿': 'Keling, tanlov haqida sizga batafsil maʻlumotlar berib oʻtaman:'},
             'resume_question': {
                 'Русский 🇷🇺': 'Я помогу тебе заполнить анкету для участия в отборе на вакансию оператора контакт-центра Билайн!\n\n Я задам 10 вопросов - это займёт не больше 5 минут 😉\n\nНачнём?\nЖми 👉 «Продолжить»',
                 'Oʻzbek tili 🇺🇿': 'Men sizga “Beeline” aloqa markazi operatori bo‘sh ish o‘rinlariga tanlovda ishtirok etish uchun anketani to‘ldirishda yordam beraman!\n\n Men 10-ta savol beraman - bu 5 daqiqadan kam vaqtni oladi 😉\n\nBoshlaymizmi?\n"Davom etish" 👈🏻 tugmasini bosing'},
             'salom': {'Русский 🇷🇺': 'Привет 👋', 'Oʻzbek tili 🇺🇿': 'Salom 👋'},
             'day': {'Русский 🇷🇺': 'День', 'Oʻzbek tili 🇺🇿': 'kun'},
             'month': {'Русский 🇷🇺': 'Месяц', 'Oʻzbek tili 🇺🇿': 'oy'},
             'year': {'Русский 🇷🇺': 'Год', 'Oʻzbek tili 🇺🇿': 'yil'},
             'choose_day': {'Русский 🇷🇺': 'Выбери день', 'Oʻzbek tili 🇺🇿': 'Kunni tanlang'},
             'choose_month': {'Русский 🇷🇺': 'Выбери месяц', 'Oʻzbek tili 🇺🇿': 'Oyni tanlang'},
             'choose_year': {'Русский 🇷🇺': 'Выбери год', 'Oʻzbek tili 🇺🇿': 'Yilni tanlang'},
             'january': {'Русский 🇷🇺': 'Январь', 'Oʻzbek tili 🇺🇿': 'Yanvar'},
             'february': {'Русский 🇷🇺': 'Февраль', 'Oʻzbek tili 🇺🇿': 'Fevral'},
             'march': {'Русский 🇷🇺': 'Март', 'Oʻzbek tili 🇺🇿': 'Mart'},
             'april': {'Русский 🇷🇺': 'Апрель', 'Oʻzbek tili 🇺🇿': 'Aprel'},
             'may': {'Русский 🇷🇺': 'Май', 'Oʻzbek tili 🇺🇿': 'May'},
             'june': {'Русский 🇷🇺': 'Июнь', 'Oʻzbek tili 🇺🇿': 'Iyun'},
             'july': {'Русский 🇷🇺': 'Июль', 'Oʻzbek tili 🇺🇿': 'Iyul'},
             'august': {'Русский 🇷🇺': 'Август', 'Oʻzbek tili 🇺🇿': 'Avgust'},
             'september': {'Русский 🇷🇺': 'Сентябрь', 'Oʻzbek tili 🇺🇿': 'Sentyabr'},
             'october': {'Русский 🇷🇺': 'Октябрь', 'Oʻzbek tili 🇺🇿': 'Oktyabr'},
             'november': {'Русский 🇷🇺': 'Ноябрь', 'Oʻzbek tili 🇺🇿': 'Noyabr'},
             'december': {'Русский 🇷🇺': 'Декабрь', 'Oʻzbek tili 🇺🇿': 'Dekabr'},
             'send': {'Русский 🇷🇺': 'Отправить', 'Oʻzbek tili 🇺🇿': 'Yuborish'},
             'd_not_choosen': {'Русский 🇷🇺': ' Ты не выбрал день', 'Oʻzbek tili 🇺🇿': 'Siz kunni tanlamadingiz'},
             'm_not_choosen': {'Русский 🇷🇺': 'Ты не выбрал месяц', 'Oʻzbek tili 🇺🇿': 'Siz oyni tanlamadingiz'},
             'y_not_choosen': {'Русский 🇷🇺': 'Ты не выбрал год', 'Oʻzbek tili 🇺🇿': 'Siz yilni tanlamadingiz'},
             'd/y_not_choosen': {'Русский 🇷🇺': 'Ты не выбрал день и год',
                                 'Oʻzbek tili 🇺🇿': 'Siz kun va yilni tanlamadingiz'},
             'd/m_not_choosen': {'Русский 🇷🇺': 'Ты не выбрал день и месяц',
                                 'Oʻzbek tili 🇺🇿': 'Siz kun va oyni tanlamadingiz'},
             'm/y_not_choosen': {'Русский 🇷🇺': 'Ты не выбрал месяц и год',
                                 'Oʻzbek tili 🇺🇿': 'Siz oy va yilni tanlamadingiz'},
             'data_ne_vibrana': {'Русский 🇷🇺': 'Ты не выбрал дату', 'Oʻzbek tili 🇺🇿': 'Siz sanani tanlamadingiz'},
             'data_not_exist': {'Русский 🇷🇺': 'Такой даты не существует', 'Oʻzbek tili 🇺🇿': 'Bunday sana yoʻq'},
             'rejection': {'Русский 🇷🇺': 'Ты отказался от составления резюме',
                           'Oʻzbek tili 🇺🇿': 'Siz anketa toʻldirishdan voz kechdingiz'},

             'less_than_18': {'Русский 🇷🇺': 'Спасибо тебе за ответы на вопросы!\nЕсли твоя кандидатура подойдёт после рассмотрения, то мы тебе перезвоним.',
                           'Oʻzbek tili 🇺🇿': 'Savollarga javob berganingiz uchun tashakkur!\nAgar sizning nomzodingiz mos kelsa, biz sizga qoʻngʻiroq qilamiz.'},
             'kirill_name': {'Русский 🇷🇺': 'Введи имя на кириллице!',
                           'Oʻzbek tili 🇺🇿': 'Ismni kirill alifbosida kiriting!'},
             'kirill_surname': {'Русский 🇷🇺': 'Введи фамилию на кириллице!',
                           'Oʻzbek tili 🇺🇿': 'Familiyangizni kirill alifbosida kiriting!'},             
             'otkaz': {'Русский 🇷🇺': 'Жаль :(\nПожалуйста, поделись с нами причинами твоего отказа.\nТы нам очень поможешь улучшить наши процессы рекрутинга',
                           'Oʻzbek tili 🇺🇿': 'Afsus :(\nIltimos, rad etishingiz sabablarini biz bilan baham koʻring.\nIshga qabul qilish jarayonlarimizni yaxshilashda bizga katta yordam berasiz'},
             
             'time_for_call': {'Русский 🇷🇺': 'Подскажи нам, в какие дни недели и в какое время тебе будет комфортно пройти телефонное интервью',
                           'Oʻzbek tili 🇺🇿': 'Haftaning qaysi kunlari va qaysi vaqtda telefon orqali suhbat oʻtkazish siz uchun qulay boʻlishini ayting'},
             'thanks_for_comment': {'Русский 🇷🇺': 'Спасибо вам за ваши комментарии!\nВаша обратная связь очень ценна для будущих соискателей 🙏',
                           'Oʻzbek tili 🇺🇿': 'Qo’shimcha ma’lumotlaringiz uchun minnatdormiz! Bir to’g’rimizdagi fikrlaringiz uchun rahmat va bu fikrlaringiz sizdan keyingi ishlovchilar uchun juda muhim ma’lumot🙏'}  
             
             

             }


class User:
    def __init__(self):
        self.cause = None
        self.name = None
        self.surname = None
        self.day = '-'
        self.month = '-'
        self.year = '-'
        self.number = None
        self.town = None
        self.district = None
        self.town_and_district = None
        self.education = None
        self.uz_language = None
        self.ru_language = None
        self.en_language = None
        self.work = None
        self.work_experience = None
        self.time_for_call = None
        
        




# markups for calendar


markup_calendar_day = types.InlineKeyboardMarkup(row_width=7)
item1 = types.InlineKeyboardButton('1', callback_data='1')
item2 = types.InlineKeyboardButton('2', callback_data='2')
item3 = types.InlineKeyboardButton('3', callback_data='3')
item4 = types.InlineKeyboardButton('4', callback_data='4')
item5 = types.InlineKeyboardButton('5', callback_data='5')
item6 = types.InlineKeyboardButton('6', callback_data='6')
item7 = types.InlineKeyboardButton('7', callback_data='7')
item8 = types.InlineKeyboardButton('8', callback_data='8')
item9 = types.InlineKeyboardButton('9', callback_data='9')
item10 = types.InlineKeyboardButton('10', callback_data='10')
item11 = types.InlineKeyboardButton('11', callback_data='11')
item12 = types.InlineKeyboardButton('12', callback_data='12')
item13 = types.InlineKeyboardButton('13', callback_data='13')
item14 = types.InlineKeyboardButton('14', callback_data='14')
item15 = types.InlineKeyboardButton('15', callback_data='15')
item16 = types.InlineKeyboardButton('16', callback_data='16')
item17 = types.InlineKeyboardButton('17', callback_data='17')
item18 = types.InlineKeyboardButton('18', callback_data='18')
item19 = types.InlineKeyboardButton('19', callback_data='19')
item20 = types.InlineKeyboardButton('20', callback_data='20')
item21 = types.InlineKeyboardButton('21', callback_data='21')
item22 = types.InlineKeyboardButton('22', callback_data='22')
item23 = types.InlineKeyboardButton('23', callback_data='23')
item24 = types.InlineKeyboardButton('24', callback_data='24')
item25 = types.InlineKeyboardButton('25', callback_data='25')
item26 = types.InlineKeyboardButton('26', callback_data='26')
item27 = types.InlineKeyboardButton('27', callback_data='27')
item28 = types.InlineKeyboardButton('28', callback_data='28')
item29 = types.InlineKeyboardButton('29', callback_data='29')
item30 = types.InlineKeyboardButton('30', callback_data='30')
item31 = types.InlineKeyboardButton('31', callback_data='31')
markup_calendar_day.add(item1, item2, item3, item4, item5, item6, item7, item8, item9, item10, item11, item12, item13,
                        item14, item15, item16, item17, item18, item19, item20, item21, item22, item23, item24, item25,
                        item26, item27, item28, item29, item30, item31)

markup_calendar_year = types.InlineKeyboardMarkup(row_width=5)
item1 = types.InlineKeyboardButton('1970', callback_data='1970')
item2 = types.InlineKeyboardButton('1971', callback_data='1971')
item3 = types.InlineKeyboardButton('1972', callback_data='1972')
item4 = types.InlineKeyboardButton('1973', callback_data='1973')
item5 = types.InlineKeyboardButton('1974', callback_data='1974')
item6 = types.InlineKeyboardButton('1975', callback_data='1975')
item7 = types.InlineKeyboardButton('1976', callback_data='1976')
item8 = types.InlineKeyboardButton('1977', callback_data='1977')
item9 = types.InlineKeyboardButton('1978', callback_data='1978')
item10 = types.InlineKeyboardButton('1979', callback_data='1979')
item11 = types.InlineKeyboardButton('1980', callback_data='1980')
item12 = types.InlineKeyboardButton('1981', callback_data='1981')
item13 = types.InlineKeyboardButton('1982', callback_data='1982')
item14 = types.InlineKeyboardButton('1983', callback_data='1983')
item15 = types.InlineKeyboardButton('1984', callback_data='1984')
item16 = types.InlineKeyboardButton('1985', callback_data='1985')
item17 = types.InlineKeyboardButton('1986', callback_data='1986')
item18 = types.InlineKeyboardButton('1987', callback_data='1987')
item19 = types.InlineKeyboardButton('1988', callback_data='1988')
item20 = types.InlineKeyboardButton('1989', callback_data='1989')
item21 = types.InlineKeyboardButton('1990', callback_data='1990')
item22 = types.InlineKeyboardButton('1991', callback_data='1991')
item23 = types.InlineKeyboardButton('1992', callback_data='1992')
item24 = types.InlineKeyboardButton('1993', callback_data='1993')
item25 = types.InlineKeyboardButton('1994', callback_data='1994')
item26 = types.InlineKeyboardButton('1995', callback_data='1995')
item27 = types.InlineKeyboardButton('1996', callback_data='1996')
item28 = types.InlineKeyboardButton('1997', callback_data='1997')
item29 = types.InlineKeyboardButton('1998', callback_data='1998')
item30 = types.InlineKeyboardButton('1999', callback_data='1999')
item31 = types.InlineKeyboardButton('2000', callback_data='2000')
item32 = types.InlineKeyboardButton('2001', callback_data='2001')
item33 = types.InlineKeyboardButton('2002', callback_data='2002')
item34 = types.InlineKeyboardButton('2003', callback_data='2003')
item35 = types.InlineKeyboardButton('2004', callback_data='2004')
item36 = types.InlineKeyboardButton('2005', callback_data='2005')
item37 = types.InlineKeyboardButton('2006', callback_data='2006')
item38 = types.InlineKeyboardButton('2007', callback_data='2007')
item39 = types.InlineKeyboardButton('2008', callback_data='2008')
item40 = types.InlineKeyboardButton('2009', callback_data='2009')
markup_calendar_year.add(item1, item2, item3, item4, item5, item6, item7, item8, item9, item10, item11, item12, item13,
                         item14, item15, item16, item17, item18, item19, item20, item21, item22, item23, item24, item25,
                         item26, item27, item28, item29, item30, item31, item32, item33, item34, item35, item36, item37,
                         item38, item39, item40)


@bot.message_handler(commands=['start'])
def process_start(message):
    between_language_and_about_resume(message)


@bot.message_handler(content_types=['text'])
def checker(message):
    print(message.text)
    print("checker")
    if (message.text == '/start'):
        print("in if")
        process_start(message)
        return
    elif (message.text == 'Начать сначала'):
        process_start(message)
        return
    else:
        print("in else")
        bot.reply_to(message, "Выбери вариант кнопкой")


@bot.message_handler(content_types=['text'])
def between_language_and_about_resume(message):
    user = user_dict[message.chat.id]
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn = types.KeyboardButton('Начать сначала')
    markup.row(btn)
    bot.send_message(message.chat.id, "Привет 👋", reply_markup=markup)
    ask_about_resume(message)
    # bot.register_next_step_handler(message, ask_about_resume)


@bot.message_handler(content_types=['text'])
def ask_about_resume(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]

    markup_resume = types.InlineKeyboardMarkup(row_width=2)
    item1 = types.InlineKeyboardButton('Отказаться', callback_data='Отказаться')
    item2 = types.InlineKeyboardButton('Продолжить', callback_data='Продолжить')

    markup_resume.add(item1, item2)
    bot.send_message(message.chat.id, "Я помогу тебе заполнить анкету для участия в отборе на вакансию оператора контакт-центра Билайн!\n\n Я задам 10 вопросов - это займёт не больше 5 минут 😉\n\nНачнём?\nЖми 👉 «Продолжить»", reply_markup=markup_resume)


def ask_about_resume_second(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]

    markup_resume_second = types.InlineKeyboardMarkup(row_width=1)
    item1 = types.InlineKeyboardButton('Да, я хочу в Билайн!', callback_data='Хочу_в_билайн')
    item2 = types.InlineKeyboardButton('Не интересует', callback_data='Не_интересует')
    item3 = types.InlineKeyboardButton('Назад', callback_data='Назад к предыдущему тексту')
    markup_resume_second.add(item1, item2, item3)

    bot.send_message(message.chat.id, "Давай ещё раз уточним, что тебе предстоит пройти:\n\n1. Ты будешь рассматриваться на вакансию оператора контакт-центра Билайн\n\n2. Перед тем как устроиться на работу к нам тебе предстоит пройти 3 этапа отбора: \n- телефонное собеседование\n- тестирование в нашем офисе (там ничего сложного, даже весело 🙂)\n- прикольное собеседование с 2-3 веселыми людьми\n\n3. Если проходишь все этапы - то мы тебя зачисляем на обучение в нашем офисе. \nОбучение длится 15-17 дней. Ты можешь выбрать дневную или вечернюю форму обучения.\nЕсли проходишь обучение и сдаёшь аттестацию - ты принят в штат! 🎉😁\n\nОплата труда, график работы и команда - обо всем этом расскажем и даже покажем 😎\n\nЕсли ты умеешь общаться как минимум на 2 языках (узбекский и русский) и тебе уже 18 лет, жми «Да, я хочу в Билайн!»", reply_markup=markup_resume_second)


@bot.message_handler(content_types=['text'])
def between_about_resume_second_and_number(message):
    chat_id = message.chat.id
    number = message.text
    user = user_dict[chat_id]

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn = types.KeyboardButton('Начать сначала')
    markup.row(btn)

    markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn_1 = types.KeyboardButton('Начать сначала')
    btn_2 = types.KeyboardButton('Назад')
    markup__v1.row(btn_1, btn_2)
    bot.send_message(message.chat.id, '1⃣')
    msg = bot.send_message(message.chat.id, "Укажи контактный номер, чтобы мы могли связаться с тобой", reply_markup=markup__v1)
    bot.register_next_step_handler(msg, ask_number)


@bot.message_handler(content_types=['text'])
def ask_number(message):
    try:
        chat_id = message.chat.id
        number = message.text
        user = user_dict[chat_id]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('Начать сначала')
        markup.row(btn)

        markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn_1 = types.KeyboardButton('Начать сначала')
        btn_2 = types.KeyboardButton('Назад')
        markup__v1.row(btn_1, btn_2)

        if (number == 'Начать сначала' or number == '/start'):
            process_start(message)
            return

        if (number == 'Назад'):
            chat_id = message.chat.id
            user = user_dict[chat_id]

            markup_resume_second = types.InlineKeyboardMarkup(row_width=1)
            item1 = types.InlineKeyboardButton('Да, я хочу в Билайн!', callback_data='Хочу_в_билайн')
            item2 = types.InlineKeyboardButton('Не интересует', callback_data='Не_интересует')
            item3 = types.InlineKeyboardButton('Назад', callback_data='Назад к предыдущему тексту')
            markup_resume_second.add(item1, item2, item3)
            bot.send_message(message.chat.id, "Давай ещё раз уточним, что тебе предстоит пройти:", reply_markup=markup)
            bot.send_message(message.chat.id, "1. Ты будешь рассматриваться на вакансию оператора контакт-центра Билайн\n\n2. Перед тем как устроиться на работу к нам тебе предстоит пройти 3 этапа отбора: \n- телефонное собеседование\n- тестирование в нашем офисе (там ничего сложного, даже весело 🙂)\n- прикольное собеседование с 2-3 веселыми людьми\n\n3. Если проходишь все этапы - то мы тебя зачисляем на обучение в нашем офисе. \nОбучение длится 15-17 дней. Ты можешь выбрать дневную или вечернюю форму обучения.\nЕсли проходишь обучение и сдаёшь аттестацию - ты принят в штат! 🎉😁\n\nОплата труда, график работы и команда - обо всем этом расскажем и даже покажем 😎\n\nЕсли ты умеешь общаться как минимум на 2 языках (узбекский и русский) и тебе уже 18 лет, жми «Да, я хочу в Билайн!»", reply_markup=markup_resume_second)
            return

        if not all(x.isascii() or x.isspace() or x.isalnum() for x in number):
            msg = bot.reply_to(message, "Неверный формат номера!")
            bot.register_next_step_handler(msg, ask_number)
            return

        user.number = number
        between_resume_and_name(message)

    except Exception:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        msg = bot.reply_to(message, "Неверный формат номера!")
        bot.register_next_step_handler(msg, ask_number)


def between_resume_and_name(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn = types.KeyboardButton('Начать сначала')
    markup.row(btn)

    markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn_1 = types.KeyboardButton('Начать сначала')
    btn_2 = types.KeyboardButton('Назад')
    markup__v1.row(btn_1, btn_2)
    
    bot.send_message(message.chat.id, '2⃣')
    msg = bot.send_message(message.chat.id, "Напиши своё имя", reply_markup=markup__v1)
    bot.register_next_step_handler(msg, ask_name)


@bot.message_handler(content_types=['text'])
def ask_name(message):
    try:
        chat_id = message.chat.id
        name = message.text
        user = user_dict[chat_id]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('Начать сначала')
        markup.row(btn)

        markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn_1 = types.KeyboardButton('Начать сначала')
        btn_2 = types.KeyboardButton('Назад')
        markup__v1.row(btn_1, btn_2)

        if (name == 'Начать сначала' or name == '/start'):
            process_start(message)
            return

        if (name == 'Назад'):
            chat_id = message.chat.id
            user = user_dict[chat_id]
            between_about_resume_second_and_number(message)
            return

        if not all(x.isspace() or x.isalpha() for x in name):
            msg = bot.reply_to(message, "Данные введены некорректно. Просим указать имя")
            bot.register_next_step_handler(msg, ask_name)
            return
        
        x = re.findall("[a-zA-Z]", name)
        
        if x:
            msg = bot.reply_to(message, "Введи имя на кириллице!")
            bot.register_next_step_handler(msg, ask_name)
            return
            
        
        
        
        user.name = name

        bot.send_message(message.chat.id, '3⃣')
        msg = bot.send_message(message.chat.id, "Напиши свою фамилию", reply_markup=markup__v1)
        bot.register_next_step_handler(msg, ask_surname)

    except Exception as e:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        msg = bot.reply_to(message, "Неверные данные")
        bot.register_next_step_handler(msg, ask_name)


def between_name_and_surname(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn = types.KeyboardButton('Начать сначала')
    markup.row(btn)

    markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn_1 = types.KeyboardButton('Начать сначала')
    btn_2 = types.KeyboardButton('Назад')
    markup__v1.row(btn_1, btn_2)

    bot.send_message(message.chat.id, '3⃣')
    msg = bot.send_message(message.chat.id, "Напиши свою фамилию")
    bot.register_next_step_handler(msg, ask_surname)


@bot.message_handler(content_types=['text'])
def ask_surname(message):
    try:
        chat_id = message.chat.id
        surname = message.text
        user = user_dict[chat_id]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('Начать сначала')
        markup.row(btn)

        markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn_1 = types.KeyboardButton('Начать сначала')
        btn_2 = types.KeyboardButton('Назад')
        markup__v1.row(btn_1, btn_2)

        if (surname == 'Начать сначала' or surname == '/start'):
            process_start(message)
            return

        if (surname == 'Назад'):
            chat_id = message.chat.id
            user = user_dict[chat_id]
            between_resume_and_name(message)
            return

        if not all(x.isspace() or x.isalpha() for x in surname):
            msg = bot.reply_to(message, "Данные введены некорректно. Просим указать фамилию")
            bot.register_next_step_handler(msg, ask_surname)
            return
        
        x = re.findall("[a-zA-Z]", surname)
        
        if x:
            msg = bot.reply_to(message, "Введи фамилию на кириллице!")
            bot.register_next_step_handler(msg, ask_surname)
            return
        
        user.surname = surname
        bot.send_message(message.chat.id, '4⃣', reply_markup=markup)
        between_name_and_birthday(message)

    except Exception as e:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        msg = bot.reply_to(message, "Неверные данные")
        bot.register_next_step_handler(msg, ask_surname)


def between_name_and_birthday(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn = types.KeyboardButton('Начать сначала')
    markup.row(btn)

    markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn_1 = types.KeyboardButton('Начать сначала')
    btn_2 = types.KeyboardButton('Назад')
    markup__v1.row(btn_1, btn_2)

    markup_calendar_start = types.InlineKeyboardMarkup(row_width=3)
    item1 = types.InlineKeyboardButton('День', callback_data='День')
    item2 = types.InlineKeyboardButton('Месяц', callback_data='Месяц')
    item3 = types.InlineKeyboardButton('Год', callback_data='Год')
    item4 = types.InlineKeyboardButton('Отправить', callback_data='Отправить')
    item5 = types.InlineKeyboardButton('Назад', callback_data='bck_to_surname')
    markup_calendar_start.add(item1, item2, item3, item4, item5)
    bot.send_message(message.chat.id, "Дата твоего рождения", reply_markup=markup_calendar_start)


def ask_town(message):
    try:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        markup_town = types.InlineKeyboardMarkup(row_width=2)
        item1 = types.InlineKeyboardButton('Ташкент', callback_data='Ташкент')
        item2 = types.InlineKeyboardButton('Другой город', callback_data='Другой город')
        item3 = types.InlineKeyboardButton('Назад', callback_data='back_to_birthday')
        markup_town.add(item1, item2, item3)
        bot.send_message(message.chat.id, "Откуда ты?", reply_markup=markup_town)
    except Exception:
        msg = bot.reply_to(message, 'Неверные данные!')
        bot.register_next_step_handler(msg, ask_town)


def choose_district(message):
    try:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        markup_regions = types.InlineKeyboardMarkup(row_width=3)
        item1 = types.InlineKeyboardButton('Алмазар', callback_data='Олмазарский')
        item2 = types.InlineKeyboardButton('Бектемир', callback_data='Бектемирский')
        item3 = types.InlineKeyboardButton('Мирабад', callback_data='Мирабадский')
        item4 = types.InlineKeyboardButton('Мирзо-Улугбек', callback_data='Мирзо-Улугбекский')
        item5 = types.InlineKeyboardButton('Сергели', callback_data='Сергелинский')
        item6 = types.InlineKeyboardButton('Чиланзар', callback_data='Чиланзарский')
        item7 = types.InlineKeyboardButton('Шайхантаур', callback_data='Шайхантаурский')
        item8 = types.InlineKeyboardButton('Юнусабад', callback_data='Юнусабадский')
        item9 = types.InlineKeyboardButton('Яккасарай', callback_data='Яккасарайский')
        item10 = types.InlineKeyboardButton('Яшнабад', callback_data='Яшнабадский')
        item11 = types.InlineKeyboardButton('Учтепа', callback_data='Учтепинский')
        item12 = types.InlineKeyboardButton('Янгихаёт', callback_data='Янгихаётский')
        item13 = types.InlineKeyboardButton('Назад', callback_data='back_to_town')
        markup_regions.add(item1, item2, item3, item4, item5, item6, item7, item8, item9, item10, item11, item12, item13)
        bot.send_message(message.chat.id, "Выбери район", reply_markup=markup_regions)
    except Exception:
        msg = bot.reply_to(message, 'Неверные данные!')
        bot.register_next_step_handler(msg, ask_town)


def between_ask_town_and_ask_town_and_district(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]

    markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn_1 = types.KeyboardButton('Начать сначала')
    btn_2 = types.KeyboardButton('Назад')
    markup__v1.row(btn_1, btn_2)

    msg = bot.send_message(message.chat.id, "Напиши регион (город) и район", reply_markup=markup__v1)
    bot.register_next_step_handler(msg, ask_town_and_district)


@bot.message_handler(content_types=['text'])
def ask_town_and_district(message):
    try:
        chat_id = message.chat.id
        town_and_district = message.text
        user = user_dict[chat_id]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('Начать сначала')
        markup.row(btn)

        markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn_1 = types.KeyboardButton('Начать сначала')
        btn_2 = types.KeyboardButton('Назад')
        markup__v1.row(btn_1, btn_2)

        if (town_and_district == 'Назад'):
            chat_id = message.chat.id
            user = user_dict[chat_id]
            bot.send_message(message.chat.id, '5⃣', reply_markup=markup)
            ask_town(message)
            return
        if (town_and_district == 'Начать сначала' or town_and_district == '/start'):
            process_start(message)
            return
        if not all(x.isascii() or x.isspace() or x.isalnum() for x in town_and_district):
            msg = bot.reply_to(message.chat.id, "Название города должно состоять из букв и может быть несколькими словами")
            bot.register_next_step_handler(msg, ask_town_and_district)
            return

        user.town_and_district = town_and_district
        

        education_1(message)


    except Exception:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        msg = bot.reply_to(message, "Неверные данные")
        bot.register_next_step_handler(msg, ask_town_and_district)


@bot.message_handler(content_types=['text'])
def education_1(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]
    print(user.town_and_district)
    markup1 = types.InlineKeyboardMarkup(row_width=1)
    item1 = types.InlineKeyboardButton('Высшее', callback_data='Высшее')
    item2 = types.InlineKeyboardButton('Неполное высшее', callback_data='Неполное высшее')
    item3 = types.InlineKeyboardButton('Среднее', callback_data='Среднее')
    item4 = types.InlineKeyboardButton('Неполное среднее', callback_data='Неполное среднее')
    item5 = types.InlineKeyboardButton('Среднее специальное', callback_data='Среднее специальное')
    item6 = types.InlineKeyboardButton('Назад', callback_data='back_to_town')
    markup1.add(item1, item2, item3, item4, item5, item6)
    bot.send_message(message.chat.id, "6⃣")
    bot.send_message(message.chat.id, "Укажи уровень образования", reply_markup=markup1)


@bot.message_handler(content_types=['text'])
def uzb_language(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]
    markup2 = types.InlineKeyboardMarkup(row_width=1)
    item1 = types.InlineKeyboardButton('Владею как родным языком :)', callback_data='Отлично')
    item2 = types.InlineKeyboardButton('Могу вести обычные беседы свободно и уверенно', callback_data='Хорошо')
    item3 = types.InlineKeyboardButton('Понимаю язык, но не могу свободно общаться', callback_data='Удовлетворительно')
    item4 = types.InlineKeyboardButton('Не владею языком', callback_data='Не владею узбекским языком')
    item5 = types.InlineKeyboardButton('Назад', callback_data='bck_ru')
    markup2.add(item1, item2, item3, item4, item5)
    bot.send_message(message.chat.id, "8⃣")
    bot.send_message(message.chat.id, "Уровень владения Узбекским языком", reply_markup=markup2)


@bot.message_handler(content_types=['text'])
def rus_language(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]
    markup3 = types.InlineKeyboardMarkup(row_width=1)
    item1 = types.InlineKeyboardButton('Владею как родным языком :)', callback_data='Отлично знаю')
    item2 = types.InlineKeyboardButton('Могу вести обычные беседы свободно и уверенно', callback_data='Хорошо знаю')
    item3 = types.InlineKeyboardButton('Понимаю язык, но не могу свободно общаться', callback_data='Удовлетворительно знаю')
    item4 = types.InlineKeyboardButton('Не владею языком', callback_data='Не владею русским языком')
    item5 = types.InlineKeyboardButton('Назад', callback_data='bck_edu')
    markup3.add(item1, item2, item3, item4, item5)
    bot.send_message(message.chat.id, "7⃣")
    bot.send_message(message.chat.id, "Уровень владения Русским языком", reply_markup=markup3)


@bot.message_handler(content_types=['text'])
def english_language(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]
    markup4 = types.InlineKeyboardMarkup(row_width=1)
    item1 = types.InlineKeyboardButton('Владею как родным языком :)', callback_data='Отлично владею')
    item2 = types.InlineKeyboardButton('Могу вести обычные беседы свободно и уверенно', callback_data='Хорошо владею')
    item3 = types.InlineKeyboardButton('Понимаю язык, но не могу свободно общаться', callback_data='Удовлетворительно владею')
    item4 = types.InlineKeyboardButton('Не владею языком', callback_data='Не владею английским языком')
    item5 = types.InlineKeyboardButton('Назад', callback_data='bck_uz')
    markup4.add(item1, item2, item3, item4, item5)
    bot.send_message(message.chat.id, "9⃣")
    bot.send_message(message.chat.id, "Уровень владения Английским языком", reply_markup=markup4)


@bot.message_handler(content_types=['text'])
def about_work(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]
    markup_o = types.InlineKeyboardMarkup(row_width=2)
    item1 = types.InlineKeyboardButton('Есть', callback_data='да')
    item2 = types.InlineKeyboardButton('Нет', callback_data='нет')
    item3 = types.InlineKeyboardButton('Назад', callback_data='bck_eng')
    markup_o.row(item1, item2)
    markup_o.row(item3)
    bot.send_message(message.chat.id, "🔟")
    bot.send_message(message.chat.id, "Есть ли у тебя опыт работы?", reply_markup=markup_o)


@bot.message_handler(content_types=['text'])
def say_experience(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]
    markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn_1 = types.KeyboardButton('Начать сначала')
    btn_2 = types.KeyboardButton('Назад')
    markup__v1.row(btn_1, btn_2)
    msg = bot.send_message(message.chat.id, "Опиши последний опыт работы коротко\n\n- Кем ты работал? \n- В какой организации? \n- Период работы", reply_markup=markup__v1)
    bot.register_next_step_handler(msg, ask_work_experience)


@bot.message_handler(content_types=['text'])
def ask_work_experience(message):
    try:
        chat_id = message.chat.id
        work_experience = message.text
        user = user_dict[chat_id]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('Начать сначала')
        markup.row(btn)

        markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn_1 = types.KeyboardButton('Начать сначала')
        btn_2 = types.KeyboardButton('Назад')
        markup__v1.row(btn_1, btn_2)

        if (work_experience == 'Назад'):
            chat_id = message.chat.id
            user = user_dict[chat_id]

            markup_o = types.InlineKeyboardMarkup(row_width=2)
            item1 = types.InlineKeyboardButton('Есть', callback_data='да')
            item2 = types.InlineKeyboardButton('Нет', callback_data='нет')
            item3 = types.InlineKeyboardButton('Назад', callback_data='bck_eng')
            markup_o.row(item1, item2)
            markup_o.row(item3)
            bot.send_message(message.chat.id, "🔟", reply_markup=markup)
            bot.send_message(message.chat.id, "Есть ли у тебя опыт работы? (неважно официальный или неофициальный)", reply_markup=markup_o)
            return
        if (work_experience == 'Начать сначала' or work_experience == '/start'):
            process_start(message)
            return
        if not all(x.isascii() or x.isspace() or x.isalnum() for x in work_experience):
            msg = bot.reply_to(message, "Неверные данные")
            bot.register_next_step_handler(msg, ask_work_experience)
            return
        user.work_experience = work_experience
   
        msg = bot.send_message(message.chat.id, "Подскажи нам, в какие дни недели и в какое время тебе будет комфортно пройти телефонное интервью")
        bot.register_next_step_handler(msg, ask_time_for_call)

       


    except Exception:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        msg = bot.reply_to(message, "Неверные данные")
        bot.register_next_step_handler(msg, ask_work_experience)
        
        
@bot.message_handler(content_types=['text'])
def ask_time_for_call(message):
    try:
        chat_id = message.chat.id
        time_for_call = message.text
        user = user_dict[chat_id]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('Начать сначала')
        markup.row(btn)

        markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn_1 = types.KeyboardButton('Начать сначала')
        btn_2 = types.KeyboardButton('Назад')
        markup__v1.row(btn_1, btn_2)

        if (time_for_call == 'Назад'):
            chat_id = message.chat.id
            user = user_dict[chat_id]
            
            markup_o = types.InlineKeyboardMarkup(row_width=2)
            item1 = types.InlineKeyboardButton('Есть', callback_data='да')
            item2 = types.InlineKeyboardButton('Нет', callback_data='нет')
            item3 = types.InlineKeyboardButton('Назад', callback_data='bck_eng')
            markup_o.row(item1, item2)
            markup_o.row(item3)
            
            bot.send_message(message.chat.id, "🔟", reply_markup=markup__v1)
            bot.send_message(message.chat.id, "Есть ли у тебя опыт работы? (неважно официальный или неофициальный)", reply_markup=markup_o)
            return
        if (time_for_call == 'Начать сначала' or time_for_call == '/start'):
            process_start(message)
            return
        if not all(x.isascii() or x.isspace() or x.isalnum() for x in time_for_call):
            msg = bot.reply_to(message, "Неверные данные")
            bot.register_next_step_handler(msg, ask_time_for_call)
            return
        user.time_for_call = time_for_call
        msg = bot.send_message(message.chat.id, "Супер!👍\nСпасибо за прохождение опроса!\n\nТеперь наш рекрутер оценит твой профиль и созвонится с тобой для проведения первого интервью!")

        now = datetime.now()
        response_date = now.strftime("%d.%m.%Y %H:%M:%S")

        birthday = user.day + "." + str(user.month).replace(" ", "") + "." + user.year
   
        if(user.town == None):
            town_and_district = user.town_and_district
        else:
            town_and_district = user.town + "/" + user.district
            
            
        wb = load_workbook(filename)
        ws = wb['Лист1']
        ws.append([response_date, user.cause, user.surname, user.name, user.number, birthday, town_and_district, user.education, user.uz_language, user.ru_language, user.en_language,
                   user.work_experience, user.time_for_call])
        wb.save(filename)
        print("saved 1")
        wb.close()

        bot.send_message(message.chat.id, "Несколько советов  к телефонному интервью:\n\nУбедись, что тебе ничего не будет мешать и ты не будешь отвлекаться на посторонние дела во время телефонного интервью.\n\nЕсли тебе будет неудобно говорить - не стесняйся попросить перенести звонок.\n\nСписок примерных вопросов:\n1. Коротко расскажи о себе\n2. Подробнее расскажи о причинах того, почему ты хочешь работать в Билайн сотрудником Контакт-центра\n3. Что ты ожидаешь получить от работы в Билайн\n\nУдачи :)")

        markup_start = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('/start')
        markup_start.row(btn)

        bot.send_message(message.chat.id, "Если хочешь пройти опрос заново нажми на кнопку: /start", reply_markup=markup_start)


    except Exception:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        msg = bot.reply_to(message, "Неверные данные")
        bot.register_next_step_handler(msg, ask_time_for_call)     


@bot.message_handler(content_types=['text'])
def say_thanks(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]
    msg = bot.send_message(message.chat.id, "Супер!👍\nСпасибо за прохождение опроса!\n\nТеперь наш рекрутер оценит твой профиль и созвонится с тобой для проведения первого интервью!")

    bot.send_message(message.chat.id, "Несколько советов  к телефонному интервью:\n\nУбедись, что тебе ничего не будет мешать и ты не будешь отвлекаться на посторонние дела во время телефонного интервью.\n\nЕсли тебе будет неудобно говорить - не стесняйся попросить перенести звонок.\n\nСписок примерных вопросов:\n1. Коротко расскажи о себе\n2. Подробнее расскажи о причинах того, почему ты хочешь работать в Билайн сотрудником Контакт-центра\n3. Что ты ожидаешь получить от работы в Билайн\n\nУдачи :)")

    markup_start = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn = types.KeyboardButton('/start')
    markup_start.row(btn)

    bot.send_message(message.chat.id, "Если хочешь пройти опрос заново нажми на кнопку: /start", reply_markup=markup_start)


@bot.callback_query_handler(func=lambda call: True)
def edu(call):
    message = call.message
    try:
        if call.data == 'Высшее':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Высшее", reply_markup=markup)
            education = call.data
            user.education = education
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            rus_language(message)
        if call.data == 'Неполное высшее':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Неполное высшее", reply_markup=markup)
            education = call.data
            user.education = education
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            rus_language(message)
        if call.data == 'Среднее':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Среднее", reply_markup=markup)
            education = call.data
            user.education = education
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            rus_language(message)
        if call.data == 'Неполное среднее':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Неполное среднее", reply_markup=markup)
            education = call.data
            user.education = education
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            rus_language(message)
        if call.data == 'Среднее специальное':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Среднее специальное", reply_markup=markup)
            education = call.data
            user.education = education
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            rus_language(message)

        if call.data == 'Отлично':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Владею как родным языком :)", reply_markup=markup)
            uz_language = call.data

            user.uz_language = uz_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            english_language(message)
        if call.data == 'Хорошо':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Могу вести обычные беседы свободно и уверенно", reply_markup=markup)
            uz_language = call.data

            user.uz_language = uz_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            english_language(message)
        if call.data == 'Удовлетворительно':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Понимаю язык, но не могу свободно общаться", reply_markup=markup)
            uz_language = call.data

            user.uz_language = uz_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            english_language(message)

        if call.data == 'Не владею узбекским языком':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Не владею языком", reply_markup=markup)
            uz_language = call.data

            user.uz_language = uz_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            english_language(message)

        if call.data == 'Отлично знаю':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Владею как родным языком :)", reply_markup=markup)
            ru_language = call.data

            user.ru_language = ru_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            uzb_language(message)
        if call.data == 'Хорошо знаю':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Могу вести обычные беседы свободно и уверенно", reply_markup=markup)
            ru_language = call.data

            user.ru_language = ru_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            uzb_language(message)
        if call.data == 'Удовлетворительно знаю':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Понимаю язык, но не могу свободно общаться", reply_markup=markup)
            ru_language = call.data

            user.ru_language = ru_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            uzb_language(message)

        if call.data == 'Не владею русским языком':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Не владею языком", reply_markup=markup)
            ru_language = call.data

            user.ru_language = ru_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            uzb_language(message)

        if call.data == 'Отлично владею':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Владею как родным языком :)", reply_markup=markup)
            en_language = call.data

            user.en_language = en_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            about_work(message)
        if call.data == 'Хорошо владею':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Могу вести обычные беседы свободно и уверенно", reply_markup=markup)
            en_language = call.data

            user.en_language = en_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            about_work(message)
        if call.data == 'Удовлетворительно владею':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Понимаю язык, но не могу свободно общаться", reply_markup=markup)
            en_language = call.data

            user.en_language = en_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            about_work(message)

        if call.data == 'Не владею английским языком':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Не владею языком", reply_markup=markup)
            en_language = call.data

            user.en_language = en_language
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            about_work(message)

        if call.data == 'да':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            work = call.data

            user.work = work
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            say_experience(message)

        if call.data == 'нет':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)
            
            markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn_1 = types.KeyboardButton('Начать сначала')
            btn_2 = types.KeyboardButton('Назад')
            markup__v1.row(btn_1, btn_2)

            work = call.data

            user.work = work
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            
            msg = bot.send_message(message.chat.id, "Подскажи нам, в какие дни недели и в какое время тебе будет комфортно пройти телефонное интервью", reply_markup=markup__v1 )   
            bot.register_next_step_handler(msg, ask_time_for_call)
            
        if call.data == 'bck_edu':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            bot.send_message(call.message.chat.id, "Назад")
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'bck_uz':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            bot.send_message(call.message.chat.id, "Назад")
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            uzb_language(message)

        if call.data == 'bck_ru':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            bot.send_message(call.message.chat.id, "Назад")
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            rus_language(message)

        if call.data == 'bck_eng':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            english_language(message)

        if call.data == 'back_to_town':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            user.town = 'Null'
            user.district = 'Null'
            user.town_and_district = 'Null'
            bot.send_message(call.message.chat.id, "Назад", reply_markup=markup)
            bot.send_message(message.chat.id, "5⃣")
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            ask_town(message)

        if call.data == 'back_to_birthday':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.send_message(call.message.chat.id, "Назад", reply_markup=markup)
            bot.send_message(message.chat.id, '4⃣')
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            between_name_and_birthday(message)

        if call.data == 'Ташкент':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Ташкент", reply_markup=markup)
            town = call.data
            user.town = town
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            choose_district(message)

        if call.data == 'Другой город':
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            between_ask_town_and_ask_town_and_district(message)

        if call.data == 'Олмазарский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Алмазар", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)
            
        if call.data == 'Янгихаётский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Янгихаёт", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)    

        if call.data == 'Бектемирский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Бектемир", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'Мирабадский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Мирабад", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'Мирзо-Улугбекский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Мирзо-Улугбек", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'Сергелинский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Сергели", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'Чиланзарский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Чиланзар", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'Шайхантаурский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Шайхантаур", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'Юнусабадский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Юнусабад", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)
        if call.data == 'Яккасарайский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Яккасарай", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)
        if call.data == 'Яшнабадский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Яшнабад", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'Учтепинский':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            bot.reply_to(message, "Учтепа", reply_markup=markup)
            district = call.data

            user.district = district
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            education_1(message)

        if call.data == 'Продолжить':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)
            
            wb = load_workbook(filename)
            ws = wb['Лист2']
            ws['A2'].value = ws['A2'].value+1
            wb.save(filename)
            wb.close()

            ask_about_resume_second(message)
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)

        if call.data == 'Отказаться':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            wb = load_workbook(filename)
            ws = wb['Лист2']
            ws['B2'].value = ws['B2'].value+1
            wb.save(filename)
            wb.close()
            
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
       
            say_the_cause(message)

        if call.data == 'Хочу_в_билайн':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn_1 = types.KeyboardButton('Начать сначала')
            btn_2 = types.KeyboardButton('Назад')
            markup__v1.row(btn_1, btn_2)
            
            wb = load_workbook(filename)
            ws = wb['Лист2']
            ws['C2'].value = ws['C2'].value+1
            wb.save(filename)
            wb.close()
            
            
            between_about_resume_second_and_number(message)
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)

        if call.data == 'Не_интересует':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]
            
           
            
            wb = load_workbook(filename)
            ws = wb['Лист2']
            ws['D2'].value = ws['D2'].value+1
            wb.save(filename)
            wb.close()
            
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            
            say_the_cause(message)

        if call.data == 'Назад к предыдущему тексту':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            ask_about_resume(message)
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)

        if call.data == 'День':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]
            msg = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                        text="Выбери день", parse_mode='Markdown')
            msg = bot.edit_message_reply_markup(call.from_user.id, call.message.message_id,
                                                reply_markup=markup_calendar_day)

        if call.data == 'Месяц':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]
            markup_calendar_month = types.InlineKeyboardMarkup(row_width=4)
            item1 = types.InlineKeyboardButton('Январь', callback_data='0 1')
            item2 = types.InlineKeyboardButton('Февраль', callback_data='0 2')
            item3 = types.InlineKeyboardButton('Март', callback_data='0 3')
            item4 = types.InlineKeyboardButton('Апрель', callback_data='0 4')
            item5 = types.InlineKeyboardButton('Май', callback_data='0 5')
            item6 = types.InlineKeyboardButton('Июнь', callback_data='0 6')
            item7 = types.InlineKeyboardButton('Июль', callback_data='0 7')
            item8 = types.InlineKeyboardButton('Август', callback_data='0 8')
            item9 = types.InlineKeyboardButton('Сентябрь', callback_data='0 9')
            item10 = types.InlineKeyboardButton('Октябрь', callback_data='1 0')
            item11 = types.InlineKeyboardButton('Ноябрь', callback_data='1 1')
            item12 = types.InlineKeyboardButton('Декабрь', callback_data='1 2')
            markup_calendar_month.add(item1, item2, item3, item4, item5, item6, item7, item8, item9, item10, item11,
                                      item12)
            msg = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                        text="Выбери месяц", parse_mode='Markdown')
            msg = bot.edit_message_reply_markup(call.from_user.id, call.message.message_id,
                                                reply_markup=markup_calendar_month)

        if call.data == 'Год':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]
            msg = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                        text="Выбери год", parse_mode='Markdown')
            msg = bot.edit_message_reply_markup(call.from_user.id, call.message.message_id,
                                                reply_markup=markup_calendar_year)

        if call.data == 'bck_to_surname':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup__v1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn_1 = types.KeyboardButton('Начать сначала')
            btn_2 = types.KeyboardButton('Назад')
            markup__v1.row(btn_1, btn_2)

            bot.send_message(call.message.chat.id, "Назад" reply_markup=markup__v1)
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            between_name_and_surname(message)

        # Календарные

        if call.data == '1' or call.data == '2' or call.data == '3' or call.data == '4' or call.data == '5' or call.data == '6' or call.data == '7' or call.data == '8' or call.data == '9' or call.data == '10' or call.data == '11' or call.data == '12' or call.data == '13' or call.data == '14' or call.data == '15' or call.data == '16' or call.data == '17' or call.data == '18' or call.data == '19' or call.data == '20' or call.data == '21' or call.data == '22' or call.data == '23' or call.data == '24' or call.data == '25' or call.data == '26' or call.data == '27' or call.data == '28' or call.data == '29' or call.data == '30' or call.data == '31':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            day = call.data
            user.day = day

            markup_calendar_start = types.InlineKeyboardMarkup(row_width=3)
            item1 = types.InlineKeyboardButton(user.day, callback_data='День')
            item2 = types.InlineKeyboardButton(user.month, callback_data='Месяц')
            item3 = types.InlineKeyboardButton(user.year, callback_data='Год')
            item4 = types.InlineKeyboardButton('Отправить', callback_data='Отправить')
            item5 = types.InlineKeyboardButton('Назад', callback_data='bck_to_name')
            markup_calendar_start.add(item1, item2, item3, item4, item5)
            msg = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                        text="Дата твоего рождения", parse_mode='Markdown')
            msg = bot.edit_message_reply_markup(call.from_user.id, call.message.message_id,
                                                reply_markup=markup_calendar_start)

        if call.data == '0 1' or call.data == '0 2' or call.data == '0 3' or call.data == '0 4' or call.data == '0 5' or call.data == '0 6' or call.data == '0 7' or call.data == '0 8' or call.data == '0 9' or call.data == '1 0' or call.data == '1 1' or call.data == '1 2':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            month = call.data
            user.month = month

            markup_calendar_start = types.InlineKeyboardMarkup(row_width=3)
            item1 = types.InlineKeyboardButton(user.day, callback_data='День')
            item2 = types.InlineKeyboardButton(user.month, callback_data='Месяц')
            item3 = types.InlineKeyboardButton(user.year, callback_data='Год')
            item4 = types.InlineKeyboardButton('Отправить', callback_data='Отправить')
            item5 = types.InlineKeyboardButton('Назад', callback_data='bck_to_name')
            markup_calendar_start.add(item1, item2, item3, item4, item5)
            msg = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                        text="Дата твоего рождения", parse_mode='Markdown')
            msg = bot.edit_message_reply_markup(call.from_user.id, call.message.message_id,
                                                reply_markup=markup_calendar_start)

        if call.data == '1970' or call.data == '1971' or call.data == '1972' or call.data == '1973' or call.data == '1974' or call.data == '1975' or call.data == '1976' or call.data == '1977' or call.data == '1978' or call.data == '1979' or call.data == '1980' or call.data == '1981' or call.data == '1982' or call.data == '1983' or call.data == '1984' or call.data == '1985' or call.data == '1986' or call.data == '1987' or call.data == '1988' or call.data == '1989' or call.data == '1990' or call.data == '1991' or call.data == '1992' or call.data == '1993' or call.data == '1994' or call.data == '1995' or call.data == '1996' or call.data == '1997' or call.data == '1998' or call.data == '1999' or call.data == '2000' or call.data == '2001' or call.data == '2002' or call.data == '2003' or call.data == '2004' or call.data == '2005' or call.data == '2006' or call.data == '2007' or call.data == '2008' or call.data == '2009' or call.data == '2010':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            year = call.data
            user.year = year

            markup_calendar_start = types.InlineKeyboardMarkup(row_width=3)
            item1 = types.InlineKeyboardButton(user.day, callback_data='День')
            item2 = types.InlineKeyboardButton(user.month, callback_data='Месяц')
            item3 = types.InlineKeyboardButton(user.year, callback_data='Год')
            item4 = types.InlineKeyboardButton('Отправить', callback_data='Отправить')
            item5 = types.InlineKeyboardButton('Назад', callback_data='bck_to_name')
            markup_calendar_start.add(item1, item2, item3, item4, item5)
            msg = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                        text="Дата твоего рождения", parse_mode='Markdown')
            msg = bot.edit_message_reply_markup(call.from_user.id, call.message.message_id,
                                                reply_markup=markup_calendar_start)

        if call.data == 'Отправить':
            chat_id = call.message.chat.id
            user = user_dict[chat_id]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            btn = types.KeyboardButton('Начать сначала')
            markup.row(btn)

            wihout_spaces = str(user.month).replace(" ", "")

            if user.day == '-' and user.month == '-' and user.year == '-':
                bot.send_message(message.chat.id, "Ты не выбрал дату")
            elif user.day == '-' and user.month == '-':
                bot.send_message(message.chat.id, "Ты не выбрал день и месяц")
            elif user.day == '-' and user.year == '-':
                bot.send_message(message.chat.id, "Ты не выбрал день и год")
            elif user.month == '-' and user.year == '-':
                bot.send_message(message.chat.id, "Ты не выбрал месяц и год")
            elif user.day == '-':
                bot.send_message(message.chat.id, "Ты не выбрал день")
            elif user.month == '-':
                bot.send_message(message.chat.id, "Ты не выбрал месяц")
            elif user.year == '-':
                bot.send_message(message.chat.id, "Ты не выбрал год")
            elif user.month == '0 2' and user.day == '30':
                bot.send_message(message.chat.id, "Такой даты не существует")
            elif user.month == '0 2' and user.day == '31':
                bot.send_message(message.chat.id, "Такой даты не существует")
            elif user.month == '0 4' and user.day == '31':
                bot.send_message(message.chat.id, "Такой даты не существует")
            elif user.month == '0 6' and user.day == '31':
                bot.send_message(message.chat.id, "Такой даты не существует")
            elif user.month == '0 9' and user.day == '31':
                bot.send_message(message.chat.id, "Такой даты не существует")
            elif user.month == '1 1' and user.day == '31':
                bot.send_message(message.chat.id, "Такой даты не существует")
            else:
                bot.send_message(message.chat.id, f'{user.day}.{wihout_spaces}.{user.year}', reply_markup=markup)            
                now = datetime.now()
                response_date = now.strftime("%d.%m.%Y %H:%M:%S")
                birthday = user.day + "." + str(user.month).replace(" ", "") + "." + user.year



                if(now.year - int(user.year)<18):
                    wb = load_workbook(filename)
                    ws = wb['Лист1']
                    ws.append([response_date, user.cause, user.surname, user.name, user.number, birthday])
                    #ws.cell(row = ws.max_row, column = 5).font = opx.styles.Font(color='ff0816')
                    wb.save(filename)
                    wb.close()
                    bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
                    less_18(message)
                    
                elif(now.year - int(user.year)==18):
                    if(user.month == '0 5' or user.month == '0 6' or user.month == '0 7' or user.month == '0 8' or user.month == '0 9' or user.month == '1 0' or user.month == '1 1' or user.month == '1 2'):
                        wb = load_workbook(filename)
                        ws = wb['Лист1']
                        ws.append([response_date, user.cause, user.surname, user.name, user.number, birthday])
                        #ws.cell(row = ws.max_row, column = 5).font = opx.styles.Font(color='ff0816')
                        wb.save(filename)
                        wb.close()
                        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
                        less_18(message)
                    else:
                        bot.send_message(message.chat.id, "5⃣")
                        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
                        ask_town(message)
                
                    
     
                                         
                else:
                    bot.send_message(message.chat.id, "5⃣")
                    bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
                    ask_town(message)
                    

    except Exception as e:
        bot.reply_to(message, "Пожалуйста перезапустите бот, на сервере проводились работы")




def less_18(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]

    bot.send_message(message.chat.id, "Спасибо тебе за ответы на вопросы!\nЕсли твоя кандидатура подойдёт после рассмотрения, то мы тебе перезвоним.")

    markup_start = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn = types.KeyboardButton('/start')
    markup_start.row(btn)

    bot.send_message(message.chat.id, "Если хочешь пройти опрос заново нажми на кнопку: /start", reply_markup=markup_start)
    

def say_the_cause(message):
    chat_id = message.chat.id
    user = user_dict[chat_id]
    
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn = types.KeyboardButton('Начать сначала')
    markup.row(btn)
    
    msg = bot.send_message(message.chat.id, "Жаль :(\nПожалуйста, поделись с нами причинами твоего отказа.\nТы нам очень поможешь улучшить наши процессы рекрутинга", reply_markup=markup)
    bot.register_next_step_handler(msg, say_cause_of_rejecton)    
    
    
@bot.message_handler(content_types=['text'])
def say_cause_of_rejecton(message):
    try:
        chat_id = message.chat.id
        cause = message.text
        user = user_dict[chat_id]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('Начать сначала')
        markup.row(btn)

        
        if (cause == 'Начать сначала' or cause == '/start'):
            process_start(message)
            return
        if not all(x.isascii() or x.isspace() or x.isalnum() for x in cause):
            msg = bot.reply_to(message, "Неверные данные")
            bot.register_next_step_handler(msg, say_cause_of_rejecton)
            return
        user.cause = cause
        

        now = datetime.now()
        response_date = now.strftime("%d.%m.%Y %H:%M:%S")

            
        wb = load_workbook(filename)
        ws = wb['Лист1']
        ws.append([response_date, user.cause])
        wb.save(filename)
        wb.close()

        markup_start = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        btn = types.KeyboardButton('/start')
        markup_start.row(btn)
        bot.send_message(message.chat.id, "Спасибо вам за ваши комментарии!\nВаша обратная связь очень ценна для будущих соискателей 🙏")
        bot.send_message(message.chat.id, "Если хочешь пройти опрос заново нажми на кнопку: /start", reply_markup=markup_start)


    except Exception:
        chat_id = message.chat.id
        user = user_dict[chat_id]
        msg = bot.reply_to(message, "Неверные данные")
        bot.register_next_step_handler(msg, say_cause_of_rejecton)
    



bot.enable_save_next_step_handlers(delay=2)

bot.load_next_step_handlers()

